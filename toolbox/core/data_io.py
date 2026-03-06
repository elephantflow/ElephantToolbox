from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import jsonlines
import openpyxl


class JsonToolkit:
    @staticmethod
    def read_json(path: str | Path) -> Any:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)

    @staticmethod
    def read_jsonl(path: str | Path) -> list[dict[str, Any]]:
        with open(path, "r", encoding="utf-8") as f:
            return [json.loads(line.strip("\n")) for line in f.readlines()]

    @staticmethod
    def write_jsonl(path: str | Path, rows: list[dict[str, Any]]) -> Path:
        output = Path(path)
        output.parent.mkdir(parents=True, exist_ok=True)
        with jsonlines.open(output, mode="w") as writer:
            for row in rows:
                writer.write(row)
        return output


class ExcelToolkit:
    @staticmethod
    def read_excel(file_name: str | Path, sheet_number: int):
        wb = openpyxl.load_workbook(file_name)
        return wb.worksheets[sheet_number]

    @staticmethod
    def get_max_row_column(sheet) -> tuple[int, int]:
        return sheet.max_row, sheet.max_column

    @staticmethod
    def read_single_column(sheet, column_index: int, total_row: int, begin_row: int) -> list[Any]:
        return [sheet.cell(row=i, column=column_index).value for i in range(begin_row, total_row + 1)]

    @staticmethod
    def read_single_row(sheet, row_index: int, total_column: int, begin_column: int) -> list[Any]:
        return [sheet.cell(row=row_index, column=i).value for i in range(begin_column, total_column + 1)]

    @staticmethod
    def read_tables(
        sheet,
        read_order: str,
        total_row: int,
        total_column: int,
        begin_row: int,
        begin_column: int,
    ) -> list[list[Any]]:
        tables: list[list[Any]] = []
        if read_order == "column":
            for i in range(begin_column, total_column + 1):
                tables.append([sheet.cell(row=j, column=i).value for j in range(begin_row, total_row + 1)])
        elif read_order == "row":
            for i in range(begin_row, total_row + 1):
                tables.append([sheet.cell(row=i, column=j).value for j in range(begin_column, total_column + 1)])
        else:
            raise ValueError("read_order must be 'column' or 'row'")
        return tables

    @staticmethod
    def save_excel(sheet_name: str, data: list[list[Any]], save_path: str | Path) -> Path:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        for row in data:
            ws.append(row)

        output = Path(save_path)
        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output)
        return output
